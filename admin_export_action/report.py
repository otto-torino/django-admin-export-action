# coding: utf-8

from __future__ import unicode_literals, absolute_import

import json
from collections import namedtuple
from itertools import chain
import csv
import re
from datetime import datetime

from django.http import HttpResponse, JsonResponse
from django.utils.module_loading import import_string
try:
    from django.utils.text import force_text
except:
    from django.utils.encoding import force_text
from django.template.loader import render_to_string
from django.utils import timezone

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from six import BytesIO, text_type

from .introspection import get_model_from_path_string
from .config import get_config


DisplayField = namedtuple("DisplayField", "path field")


def get_field_display_value(objects, path, raw):
    parts = path.split('__')
    model = objects.model
    i = 0
    for p in parts:
        if model._meta.get_field(p).related_model:
            model = model._meta.get_field(p).related_model
        else:
            model = model._meta.get_field(p).model

        if i == len(parts) - 1:
            if model._meta.get_field(p).choices and len(model._meta.get_field(p).choices):
                return dict(model._meta.get_field(p).choices)[raw]
            else:
                return raw
        i += 1


def get_field_verbose_name(objects, path):
    res = []
    parts = path.split('__')
    model = objects.model
    for p in parts:
        field = model._meta.get_field(p)
        if hasattr(field, 'verbose_name') and field.verbose_name:
            res.append(str(field.verbose_name))
        else:
            res.append(str(field.name))
        if model._meta.get_field(p).related_model:
            model = model._meta.get_field(p).related_model
    return ' '.join(res)


def generate_filename(title, ends_with):
    title = title.split('.')[0]
    title.replace(' ', '_')
    title += ('_' + timezone.now().strftime("%Y-%m-%d_%H%M"))
    if not title.endswith(ends_with):
        title += ends_with
    return title


def _can_change_or_view(model, user):
    """ Return True iff `user` has either change or view permission
    for `model`.
    """
    model_name = model._meta.model_name
    app_label = model._meta.app_label
    can_change = user.has_perm(app_label + '.change_' + model_name)
    can_view = user.has_perm(app_label + '.view_' + model_name)

    return can_change or can_view


def report_to_list(queryset, display_fields, user, raw_choices=False):
    """ Create list from a report with all data filtering.

    queryset: initial queryset to generate results
    display_fields: list of field references or DisplayField models
    user: requesting user

    Returns list, message in case of issues.
    """
    model_class = queryset.model
    objects = queryset
    message = ""

    if not _can_change_or_view(model_class, user):
        return [], 'Permission Denied'

    # Convert list of strings to DisplayField objects.
    new_display_fields = []

    for display_field in display_fields:
        field_list = display_field.split('__')
        field = field_list[-1]
        path = '__'.join(field_list[:-1])

        if path:
            path += '__'  # Legacy format to append a __ here.

        df = DisplayField(path, field)
        new_display_fields.append(df)

    display_fields = new_display_fields

    # Display Values
    display_field_paths = []

    for i, display_field in enumerate(display_fields):
        model = get_model_from_path_string(model_class, display_field.path)

        if not model or _can_change_or_view(model, user):
            display_field_key = display_field.path + display_field.field

            display_field_paths.append(display_field_key)

        else:
            message += 'Error: Permission denied on access to {0}.'.format(
                display_field.name
            )

    values_list = objects.values_list(*display_field_paths)
    values_and_properties_list = [list(row) for row in values_list]

    if not raw_choices:
        return list(map(
            lambda record: list(map(lambda p, v: get_field_display_value(objects, p, v), display_field_paths, record)),
            values_and_properties_list)
        ), message
    else:
        return values_and_properties_list, message


def build_sheet(data, ws, sheet_name='report', header=None, widths=None):
    first_row = 1
    column_base = 1

    func = None
    if get_config('VALUE_TO_XLSX_CELL') is not None:
        try:
            func = import_string(get_config('VALUE_TO_XLSX_CELL'))
        except Exception as e:
            pass

    ws.title = re.sub(r'\W+', '', sheet_name)[:30]
    if header:
        for i, header_cell in enumerate(header):
            cell = ws.cell(row=first_row, column=i + column_base)
            cell.value = header_cell
            cell.font = Font(bold=True)
            if widths:
                ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

    for row in data:
        for i in range(len(row)):
            item = row[i]
            # If item is a regular string
            success = False
            if func:
                success, res = func(item)
            if success:
                row[i] = res
            else:
                if isinstance(item, datetime):
                    row[i] = item.replace(tzinfo=None)
                elif isinstance(item, str):
                    # Change it to a unicode string
                    try:
                        row[i] = text_type(item)
                    except UnicodeDecodeError:
                        row[i] = text_type(item.decode('utf-8', 'ignore'))
                elif type(item) is dict:
                    row[i] = text_type(item)
                elif type(item).__name__ == 'UUID' or type(item).__name__ == '__proxy__':
                    row[i] = str(item)
                elif type(item).__name__ == 'list':
                    row[i] = json.dumps(item)
        try:
            ws.append(row)
        except ValueError as e:
            ws.append([str(e)])
        except:
            ws.append(['Unknown Error'])


def list_to_workbook(data, title='report', header=None, widths=None):
    """ Create just a openpxl workbook from a list of data """
    wb = Workbook()
    title = re.sub(r'\W+', '', title)[:30]

    if isinstance(data, dict):
        i = 0
        for sheet_name, sheet_data in data.items():
            if i > 0:
                wb.create_sheet()
            ws = wb.worksheets[i]
            build_sheet(
                sheet_data, ws, sheet_name=sheet_name, header=header)
            i += 1
    else:
        ws = wb.worksheets[0]
        build_sheet(data, ws, header=header, widths=widths)
    return wb


def build_xlsx_response(wb, title="report"):
    """ Take a workbook and return a xlsx file response """
    title = generate_filename(title, '.xlsx')
    myfile = BytesIO()
    myfile.write(save_virtual_workbook(wb))
    response = HttpResponse(
        myfile.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % title
    response['Content-Length'] = myfile.tell()
    return response


def list_to_xlsx_response(data, title='report', header=None,
                          widths=None):
    """ Make 2D list into a xlsx response for download
    data can be a 2d array or a dict of 2d arrays
    like {'sheet_1': [['A1', 'B1']]}
    """

    wb = list_to_workbook(data, title, header, widths)
    return build_xlsx_response(wb, title=title)


def list_to_csv_response(data, title='report', header=None):
    """ Make 2D list into a csv response for download data.
    """
    response = HttpResponse(content_type="text/csv; charset=UTF-8")
    response['Content-Disposition'] = 'attachment; filename="%s.csv"' % title
    cw = csv.writer(response)

    for row in chain([header] if header else [], data):
        cw.writerow([force_text(s) for s in row])
    return response


def list_to_json_response(data, title='report', header=None):
    """ Make 2D list into a json response for download data.
    """
    if not header:
        res = data
    else:
        res = []

        for record in data:
            i = 0
            partial = {}
            for field in header:
                partial[field] = record[i]
                i += 1
            res.append(partial)

    response = JsonResponse(res, safe=False)
    response['Content-Disposition'] = 'attachment; filename="%s.json"' % title

    return response


def list_to_html_response(data, title='', header=None):
    html = render_to_string('export_action/report_html.html', locals())
    return HttpResponse(html)
